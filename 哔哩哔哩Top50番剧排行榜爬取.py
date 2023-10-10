import requests
from bs4 import BeautifulSoup
import re
from openpyxl import load_workbook
from openpyxl import Workbook
def form(value):
    excel_path = "B站番剧排行榜.xlsx"

    wb = Workbook(excel_path)
    wb.save(excel_path)

    wb = load_workbook(excel_path)
    wb.create_sheet("B站番剧排行榜")
    ws = wb.active

    for row in range(1,len(value)):
        for column in range(1,4):
            ws.cell(row,column).value = value[row-1][column-1]

    wb.save(excel_path)
    wb.close()
    print("%s     保存成功！" % excel_path)
URL = 'https://www.bilibili.com/v/popular/rank/bangumi/'
resp = requests.get(URL)
#print(resp.text)
resp.encoding = 'utf-8'

value = []
main_page = BeautifulSoup(resp.text,"html.parser")
obj = re.compile(r'<div class="info">.*?class="title">(?P<name>.*?)'
                 r'</a>.*?alt="play">(?P<Playback_volume>.*?)'
                 r'</span>.*?alt="follow">(?P<Number_likes>.*?)</span>', re.S)
resule = obj.finditer(resp.text)
for a in resule:
    dic = a.groupdict()
    value.append(list(dic.values()))
form(value)